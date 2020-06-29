# -*- coding: utf-8 -*-
"""
Created on Fri Jun 26 18:16:29 2020

@author: A677720
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import random
import faker 
from faker import Faker
import datetime

fake= Faker('en_GB')

#load workbooks and datasheets
wb = load_workbook(r'C:\Users\A677720\Documents\Python_challenges\Facility_status\Facility_status_v2.xlsx')
print(wb.active)
print (wb.sheetnames)

wbr = load_workbook(r'C:\Users\A677720\Documents\Python_challenges\Facility_status\Regions2.xlsx')
print (wbr.sheetnames)


ws=wb.active
regions = wbr.active
print(regions['A1'].value)

for i in range(2, 1002):
    j = random.randint(2,532)
    k = random.randint(2,532)
    l = random.randint(2,6)
    m = random.randint(2,7)
    n = random.randint(2,5)
    o = random.randint(2,7)
    p = random.randint(2,8)
    #j = regions[]
    #k=facility
    #l=operational_status
    #m=building_status
    #n=dispensary status
    #o=facility_type

    postcode= fake.postcode()
    today = datetime.date.today()
    phone_num = fake.phone_number()
    ans=["Yes","No"]
    answer =random.choice(ans)
    if answer == "Yes":
        oxygen = random.randint(300,1000)
    else: oxygen = 0 
    patient_pop = random.randint(200,2000)
    staff_max = (patient_pop * 0.3)
    force_gen__patients = random.randint(0,20)
    name_f = fake.name_female()
    name_m = fake.name_male()
    names = [name_f, name_m]
    name = random.choice(names)
    if name == name_f:
        Prefix= "Miss"
    else: Prefix = "Mr"
    comments = fake.text()
    stock =(["FRSM masks",random.randint(0, 1000),random.randint(0,30)],["FFP3 masks",random.randint(0, 1000),random.randint(0,30)],["Gloves",random.randint(0, 1000),random.randint(0,30)],["Aprons",random.randint(0, 1000),random.randint(0,30)],["Visors",random.randint(0, 1000),random.randint(0,30)],["Biocleanse Tubs",random.randint(0, 1000),random.randint(0,30)],["Surgical Gowns",random.randint(0, 1000),random.randint(0,30)])
    stock_lvls= pd.DataFrame(stock)
    ws['A'+str(i)] = regions['B'+str(j)].value
    ws['B'+str(i)] = regions['A'+str(j)].value
    ws['C'+str(i)] = regions['C'+str(o)].value   
    ws['D'+str(i)] = (postcode)
    ws['E'+str(i)] = (today)
    ws['F'+str(i)] = (phone_num)
    ws['G'+str(i)] = regions['D'+str(l)].value
    ws['H'+str(i)] = regions['E'+str(m)].value
    ws['I'+str(i)] = (answer)
    #ws['J'+str(i)] = regions['E'+str(l)].value     REVIEW 
    ws['K'+str(i)] = regions['I'+str(n)].value
    ws['L'+str(i)] = (patient_pop)
    ws['M'+str(i)] = round(staff_max)
    ws['N'+str(i)] = round(0.3* staff_max)
    ws['O'+str(i)] = round(0.3* staff_max)
    ws['P'+str(i)] = round(0.1* staff_max)
    ws['Q'+str(i)] = round(0.15* staff_max)
    ws['R'+str(i)] = round(0.05* staff_max)
    ws['S'+str(i)] = round(0.07* staff_max)
    ws['T'+str(i)] = round(0.03* staff_max)
    ws['U'+str(i)] = (force_gen__patients)
    ws['V'+str(i)] = round(staff_max*0.20)
    ws['W'+str(i)] = random.randint(0,600)
    #ws['X'+str(i)] = regions['A'+str(j)].value     REVIEW
    ws['Y'+str(i)] = random.choice(['Yes','No','N/A'])
    ws['Z'+str(i)] = (fake.sentence())
    ws['AA'+str(i)] = (stock_lvls.iloc[0,1])
    ws['AB'+str(i)] = (stock_lvls.iloc[0,2])
    ws['AC'+str(i)] = (stock_lvls.iloc[1,1])
    ws['AD'+str(i)] = (stock_lvls.iloc[1,2])
    ws['AE'+str(i)] = (stock_lvls.iloc[2,1])
    ws['AF'+str(i)] = (stock_lvls.iloc[2,2])
    ws['AG'+str(i)] = (stock_lvls.iloc[3,1])
    ws['AH'+str(i)] = (stock_lvls.iloc[3,2])
    ws['AI'+str(i)] = (stock_lvls.iloc[4,1])
    ws['AJ'+str(i)] = (stock_lvls.iloc[4,2])
    ws['AK'+str(i)] = (stock_lvls.iloc[5,1])
    ws['AL'+str(i)] = (stock_lvls.iloc[5,2])
    #ws['AM'+str(i)] = (stock_lvls.iloc[p,0])
    ws['AM'+str(i)] = regions['J'+str(p)].value
    ws['AN'+str(i)] = (stock_lvls.iloc[6,1])
    ws['AO'+str(i)] = (stock_lvls.iloc[6,2])
    ws['AP'+str(i)] = (answer)
    ws['AQ'+str(i)] = (oxygen)
    ws['AR'+str(i)] = random.choice(['Yes','No'])
    ws['AS'+str(i)] = (stock_lvls.iloc[4,1])
    ws['AT'+str(i)] = (Prefix)
    ws['AT'+str(i)] = (Prefix) 
    ws['AU'+str(i)] = (comments) 
wb.save(r'C:\Users\A677720\Documents\Python_challenges\Facility_status\Facility_status_v2.xlsx')  